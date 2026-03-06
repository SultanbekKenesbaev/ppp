import csv
import io
import re
import secrets
from datetime import datetime, timedelta
from pathlib import Path

from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import current_user, login_required
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload
from unidecode import unidecode

from . import db
from .models import District, Mahalla, Street, User, WorkerAssignment
from .utils import hash_password, require_role

bp = Blueprint("admin", __name__)

ALLOWED_IMPORT_EXTENSIONS = {".xlsx", ".xls", ".csv"}
PROBLEM_EXPORT_TOKEN_RE = re.compile(r"^problems_[0-9]{14}_[a-f0-9]{12}$")
PROBLEM_EXPORT_TTL_HOURS = 24


def _clean_cell(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1].strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _iter_import_rows(uploaded_file, ext: str):
    uploaded_file.stream.seek(0)

    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(filename=uploaded_file.stream, read_only=True, data_only=True)
        sheet = wb.active
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            yield row_num, list(row or [])
        return

    if ext == ".xls":
        import xlrd

        data = uploaded_file.stream.read()
        book = xlrd.open_workbook(file_contents=data)
        sheet = book.sheet_by_index(0)
        for idx in range(sheet.nrows):
            yield idx + 1, sheet.row_values(idx)
        return

    if ext == ".csv":
        data = uploaded_file.stream.read()
        text = None
        for enc in ("utf-8-sig", "cp1251", "latin-1"):
            try:
                text = data.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            raise ValueError("Не удалось прочитать CSV (кодировка)")

        sample = "\n".join(text.splitlines()[:10]) or text[:2048]
        delimiter = ","
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            delimiter = dialect.delimiter
        except Exception:
            counts = {d: sample.count(d) for d in (";", ",", "\t")}
            if max(counts.values()) > 0:
                delimiter = max(counts, key=counts.get)

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        for row_num, row in enumerate(reader, start=1):
            yield row_num, row
        return

    raise ValueError("Неподдерживаемый формат файла")


def _split_fio(value: str):
    parts = [p for p in re.split(r"\s+", value.strip()) if p]
    if len(parts) < 2:
        return None
    last_name = parts[0]
    first_name = parts[1]
    middle_name = " ".join(parts[2:]) if len(parts) > 2 else ""
    return last_name, first_name, middle_name


def _split_fio_for_export(value: str):
    parts = [p for p in re.split(r"\s+", (value or "").strip()) if p]
    if len(parts) >= 2:
        return parts[0], parts[1], " ".join(parts[2:])
    if len(parts) == 1:
        return parts[0], "", ""
    return "", "", ""

def _street_base_and_num(name: str):
    cleaned = _clean_cell(name)
    match = re.match(r"^(.*)\s\(([0-9]+)\)$", cleaned)
    if match:
        base = _clean_cell(match.group(1))
        return base, int(match.group(2))
    return cleaned, None


def _normalize_street_base(name: str) -> str:
    return _clean_cell(name).casefold()


def _collect_existing_street_suffix_state(district_id: int):
    state = {}
    streets = (
        Street.query.join(Mahalla, Street.mahalla_id == Mahalla.id)
        .filter(Mahalla.district_id == district_id)
        .all()
    )

    for street in streets:
        base_name, suffix_num = _street_base_and_num(street.name)
        base_norm = _normalize_street_base(base_name)
        key = (street.mahalla_id, base_norm)
        slot = state.setdefault(key, {"has_plain": False, "max_suffix": 0, "existing_exact_names": set()})

        street_clean = _clean_cell(street.name)
        slot["existing_exact_names"].add(street_clean.casefold())
        if suffix_num is None:
            slot["has_plain"] = True
        else:
            slot["max_suffix"] = max(slot["max_suffix"], suffix_num)

    return state


def _count_import_street_duplicates(prepared_rows: list[dict]):
    counts = {}
    for row in prepared_rows:
        key = (row["mahalla_key"], row["street_base_norm"])
        counts[key] = counts.get(key, 0) + 1
    return counts



def _login_base(last_name: str, first_name: str) -> str:
    raw = f"{last_name}{first_name}"
    raw = unidecode(raw).lower()
    normalized = re.sub(r"[^a-z0-9]+", "", raw)
    return normalized or "worker"


def _next_login(base: str, existing_logins: set[str]) -> str:
    candidate = base
    suffix = 2
    while candidate.casefold() in existing_logins:
        candidate = f"{base}{suffix}"
        suffix += 1
    existing_logins.add(candidate.casefold())
    return candidate


def _problem_rows_from_report(report: dict | None):
    if not report:
        return []
    return [row for row in report.get("rows", []) if row.get("status") != "CREATED"]


def _build_import_page_context(report=None, selected_district_id=None, problem_export_token=None):
    rows = report.get("rows", []) if report else []
    problems = [row for row in rows if row.get("status") != "CREATED"]
    success_count = len(rows) - len(problems)
    return {
        "selected_district_id": selected_district_id,
        "report": report,
        "allowed_extensions": ", ".join(sorted(ALLOWED_IMPORT_EXTENSIONS)),
        "problem_counts": {
            "all": len(rows),
            "problems": len(problems),
            "success": success_count,
        },
        "problem_export_token": problem_export_token,
        "has_problems": len(problems) > 0,
    }


def _reports_dir_path() -> Path:
    reports_dir = Path(current_app.config["UPLOAD_DIR"]) / "import_reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    return reports_dir


def _cleanup_problem_exports() -> None:
    reports_dir = _reports_dir_path()
    cutoff = datetime.utcnow() - timedelta(hours=PROBLEM_EXPORT_TTL_HOURS)
    for file_path in reports_dir.glob("problems_*.xlsx"):
        try:
            modified = datetime.utcfromtimestamp(file_path.stat().st_mtime)
            if modified < cutoff:
                file_path.unlink()
        except Exception:
            pass


def _remove_problem_export_file(token: str | None) -> None:
    if not token or not PROBLEM_EXPORT_TOKEN_RE.fullmatch(token):
        return
    file_path = _reports_dir_path() / f"{token}.xlsx"
    try:
        if file_path.exists():
            file_path.unlink()
    except Exception:
        pass


def _save_problem_export_xlsx(problem_rows: list[dict]) -> str:
    from openpyxl import Workbook

    _cleanup_problem_exports()
    token = f"problems_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{secrets.token_hex(6)}"
    file_path = _reports_dir_path() / f"{token}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Проблемы"

    ws.append([
        "Фамилия",
        "Имя",
        "Отчество",
        "Улица",
        "Махалля",
        "Причина",
        "Статус",
        "Номер строки",
    ])

    for row in problem_rows:
        last_name, first_name, middle_name = _split_fio_for_export(row.get("fio", ""))
        ws.append([
            last_name,
            first_name,
            middle_name,
            row.get("street", ""),
            row.get("mahalla", ""),
            row.get("comment", ""),
            row.get("status", ""),
            row.get("row_number", ""),
        ])

    wb.save(file_path)
    return token


def _import_workers_from_file(district: District, uploaded_file):
    ext = Path(uploaded_file.filename or "").suffix.lower()
    if ext not in ALLOWED_IMPORT_EXTENSIONS:
        raise ValueError("Поддерживаются только .xlsx, .xls, .csv")

    report = {
        "filename": uploaded_file.filename or "",
        "district_name": district.name,
        "summary": {
            "processed_rows": 0,
            "created_mahallas": 0,
            "created_streets": 0,
            "created_workers": 0,
            "assigned_workers": 0,
            "skipped": 0,
            "errors": 0,
        },
        "rows": [],
    }

    mahalla_cache = {
        m.name.casefold(): m for m in Mahalla.query.filter_by(district_id=district.id).all()
    }
    street_state = _collect_existing_street_suffix_state(district.id)

    existing_logins = {
        (login or "").casefold() for (login,) in db.session.query(User.login).all()
    }

    prepared_rows = []

    for row_number, raw_row in _iter_import_rows(uploaded_file, ext):
        if row_number == 1:
            continue

        report["summary"]["processed_rows"] += 1
        row_values = list(raw_row or [])

        first_four = [
            _clean_cell(row_values[idx] if idx < len(row_values) else "")
            for idx in range(4)
        ]
        if not any(first_four):
            report["summary"]["skipped"] += 1
            report["rows"].append(
                {
                    "row_number": row_number,
                    "mahalla": "",
                    "street": "",
                    "fio": "",
                    "status": "SKIP_EMPTY",
                    "comment": "Пустая строка",
                    "login": "",
                }
            )
            continue

        mahalla_name = _clean_cell(row_values[1] if len(row_values) > 1 else "")
        street_name_raw = _clean_cell(row_values[2] if len(row_values) > 2 else "")
        worker_fio = _clean_cell(row_values[3] if len(row_values) > 3 else "")

        if not mahalla_name or not street_name_raw or not worker_fio:
            report["summary"]["errors"] += 1
            report["rows"].append(
                {
                    "row_number": row_number,
                    "mahalla": mahalla_name,
                    "street": street_name_raw,
                    "fio": worker_fio,
                    "status": "ERROR_MISSING_REQUIRED",
                    "comment": "Нужны значения в столбцах 2, 3 и 4",
                    "login": "",
                }
            )
            continue

        fio_parts = _split_fio(worker_fio)
        if not fio_parts:
            report["summary"]["errors"] += 1
            report["rows"].append(
                {
                    "row_number": row_number,
                    "mahalla": mahalla_name,
                    "street": street_name_raw,
                    "fio": worker_fio,
                    "status": "ERROR_BAD_FIO",
                    "comment": "ФИО должно содержать минимум фамилию и имя",
                    "login": "",
                }
            )
            continue

        street_base, _ = _street_base_and_num(street_name_raw)
        if not street_base:
            report["summary"]["errors"] += 1
            report["rows"].append(
                {
                    "row_number": row_number,
                    "mahalla": mahalla_name,
                    "street": street_name_raw,
                    "fio": worker_fio,
                    "status": "ERROR_MISSING_REQUIRED",
                    "comment": "Название улицы пустое",
                    "login": "",
                }
            )
            continue

        prepared_rows.append(
            {
                "row_number": row_number,
                "mahalla_name": mahalla_name,
                "mahalla_key": mahalla_name.casefold(),
                "street_base": street_base,
                "street_base_norm": _normalize_street_base(street_base),
                "worker_fio": worker_fio,
                "fio_parts": fio_parts,
            }
        )

    import_dup_counts = _count_import_street_duplicates(prepared_rows)

    for item in prepared_rows:
        row_number = item["row_number"]
        mahalla_name = item["mahalla_name"]
        worker_fio = item["worker_fio"]

        savepoint = db.session.begin_nested()
        try:
            created_mahalla = False

            mahalla_key = item["mahalla_key"]
            mahalla = mahalla_cache.get(mahalla_key)
            if not mahalla:
                mahalla = Mahalla.query.filter(
                    Mahalla.district_id == district.id,
                    func.lower(Mahalla.name) == mahalla_name.lower(),
                ).first()
            if not mahalla:
                mahalla = Mahalla(district_id=district.id, name=mahalla_name)
                db.session.add(mahalla)
                db.session.flush()
                mahalla_cache[mahalla_key] = mahalla
                created_mahalla = True

            base_name = item["street_base"]
            base_norm = item["street_base_norm"]
            state_key = (mahalla.id, base_norm)
            state = street_state.setdefault(
                state_key,
                {"has_plain": False, "max_suffix": 0, "existing_exact_names": set()},
            )

            dup_key = (mahalla_key, base_norm)
            appears_many_in_import = import_dup_counts.get(dup_key, 0) > 1
            exists_in_db_or_import = state["has_plain"] or state["max_suffix"] > 0 or bool(state["existing_exact_names"])
            needs_suffix = appears_many_in_import or exists_in_db_or_import

            suffix_num = None
            if needs_suffix:
                suffix_num = max(1, state["max_suffix"] + 1)
                candidate_name = f"{base_name} ({suffix_num})"
                while candidate_name.casefold() in state["existing_exact_names"]:
                    suffix_num += 1
                    candidate_name = f"{base_name} ({suffix_num})"
            else:
                candidate_name = base_name
                if candidate_name.casefold() in state["existing_exact_names"]:
                    suffix_num = max(1, state["max_suffix"] + 1)
                    candidate_name = f"{base_name} ({suffix_num})"
                    while candidate_name.casefold() in state["existing_exact_names"]:
                        suffix_num += 1
                        candidate_name = f"{base_name} ({suffix_num})"

            street = Street(mahalla_id=mahalla.id, name=candidate_name)
            db.session.add(street)
            db.session.flush()

            state["existing_exact_names"].add(candidate_name.casefold())
            if suffix_num is None:
                state["has_plain"] = True
            else:
                state["max_suffix"] = max(state["max_suffix"], suffix_num)

            last_name, first_name, middle_name = item["fio_parts"]
            login_base = _login_base(last_name, first_name)
            login = _next_login(login_base, existing_logins)

            worker = User(
                role="worker",
                login=login,
                password_hash=hash_password("123456"),
                first_name=first_name,
                last_name=last_name,
                middle_name=middle_name,
                note="",
            )
            db.session.add(worker)
            db.session.flush()

            db.session.add(WorkerAssignment(street_id=street.id, worker_id=worker.id))
            db.session.flush()

            savepoint.commit()

            if created_mahalla:
                report["summary"]["created_mahallas"] += 1
            report["summary"]["created_streets"] += 1
            report["summary"]["created_workers"] += 1
            report["summary"]["assigned_workers"] += 1
            report["rows"].append(
                {
                    "row_number": row_number,
                    "mahalla": mahalla_name,
                    "street": candidate_name,
                    "fio": worker_fio,
                    "status": "CREATED",
                    "comment": "Работник создан и закреплен за улицей",
                    "login": login,
                }
            )

        except Exception as exc:
            savepoint.rollback()
            report["summary"]["errors"] += 1
            report["rows"].append(
                {
                    "row_number": row_number,
                    "mahalla": mahalla_name,
                    "street": item["street_base"],
                    "fio": worker_fio,
                    "status": "ERROR_EXCEPTION",
                    "comment": str(exc).strip() or "Ошибка обработки строки",
                    "login": "",
                }
            )

    return report


@bp.get("/structure")
@login_required
def structure():
    require_role(current_user, "admin")

    per_page = request.args.get("per_page", default=50, type=int)
    per_page = max(20, min(per_page, 100))

    district_page_num = request.args.get("district_page", default=1, type=int)
    mahalla_page_num = request.args.get("mahalla_page", default=1, type=int)
    street_page_num = request.args.get("street_page", default=1, type=int)

    districts_page = (
        District.query
        .order_by(District.name.asc())
        .paginate(page=district_page_num, per_page=per_page, error_out=False)
    )
    mahallas_page = (
        Mahalla.query
        .options(joinedload(Mahalla.district))
        .order_by(Mahalla.name.asc())
        .paginate(page=mahalla_page_num, per_page=per_page, error_out=False)
    )
    streets_page = (
        Street.query
        .options(joinedload(Street.mahalla).joinedload(Mahalla.district))
        .order_by(Street.name.asc())
        .paginate(page=street_page_num, per_page=per_page, error_out=False)
    )

    districts = District.query.order_by(District.name.asc()).all()
    mahallas = Mahalla.query.options(joinedload(Mahalla.district)).order_by(Mahalla.name.asc()).all()

    return render_template(
        "admin_structure.html",
        districts=districts,
        mahallas=mahallas,
        districts_page=districts_page,
        mahallas_page=mahallas_page,
        streets_page=streets_page,
        per_page=per_page,
        district_page=district_page_num,
        mahalla_page=mahalla_page_num,
        street_page=street_page_num,
    )


@bp.get("/import")
@login_required
def import_page():
    require_role(current_user, "admin")
    districts = District.query.order_by(District.name.asc()).all()
    return render_template(
        "admin_import.html",
        districts=districts,
        **_build_import_page_context(report=None, selected_district_id=None, problem_export_token=None),
    )


@bp.post("/import")
@login_required
def import_upload():
    require_role(current_user, "admin")

    districts = District.query.order_by(District.name.asc()).all()
    district_id = request.form.get("district_id", type=int)
    uploaded_file = request.files.get("file")

    if not district_id:
        flash("Выберите район", "error")
        return redirect(url_for("admin.import_page"))

    district = District.query.get(district_id)
    if not district:
        flash("Район не найден", "error")
        return redirect(url_for("admin.import_page"))

    if not uploaded_file or not uploaded_file.filename:
        flash("Выберите файл для импорта", "error")
        return redirect(url_for("admin.import_page"))

    ext = Path(uploaded_file.filename).suffix.lower()
    if ext not in ALLOWED_IMPORT_EXTENSIONS:
        flash("Формат файла не поддерживается. Нужны .xlsx, .xls или .csv", "error")
        return redirect(url_for("admin.import_page"))

    try:
        report = _import_workers_from_file(district=district, uploaded_file=uploaded_file)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f"Импорт не выполнен: {exc}", "error")
        return redirect(url_for("admin.import_page"))

    problem_rows = _problem_rows_from_report(report)
    previous_token = session.get("last_import_problem_token")
    if previous_token:
        _remove_problem_export_file(previous_token)

    problem_export_token = None
    if problem_rows:
        try:
            problem_export_token = _save_problem_export_xlsx(problem_rows)
            session["last_import_problem_token"] = problem_export_token
        except Exception as exc:
            session.pop("last_import_problem_token", None)
            flash(f"Импорт выполнен, но экспорт проблем не создан: {exc}", "error")
    else:
        session.pop("last_import_problem_token", None)

    s = report["summary"]
    flash(
        (
            f"Импорт завершен. Обработано: {s['processed_rows']}, "
            f"создано работников: {s['created_workers']}, "
            f"пропущено: {s['skipped']}, ошибок: {s['errors']}"
        ),
        "ok",
    )

    return render_template(
        "admin_import.html",
        districts=districts,
        **_build_import_page_context(
            report=report,
            selected_district_id=district.id,
            problem_export_token=problem_export_token,
        ),
    )


@bp.get("/import/export/problems/<token>")
@login_required
def import_export_problems(token: str):
    require_role(current_user, "admin")

    if not PROBLEM_EXPORT_TOKEN_RE.fullmatch(token):
        flash("Некорректный токен экспорта", "error")
        return redirect(url_for("admin.import_page"))

    expected = session.get("last_import_problem_token")
    if not expected or expected != token:
        flash("Экспорт доступен только для последнего импорта", "error")
        return redirect(url_for("admin.import_page"))

    reports_dir = _reports_dir_path()
    file_path = reports_dir / f"{token}.xlsx"

    try:
        file_path.resolve().relative_to(reports_dir.resolve())
    except Exception:
        flash("Некорректный путь к файлу экспорта", "error")
        return redirect(url_for("admin.import_page"))

    if not file_path.exists():
        flash("Файл экспорта не найден", "error")
        return redirect(url_for("admin.import_page"))

    download_name = f"import_problems_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(file_path, as_attachment=True, download_name=download_name)


@bp.post("/district/create")
@login_required
def district_create():
    require_role(current_user, "admin")
    name = request.form.get("name", "").strip()
    if not name:
        flash("Название района обязательно", "error")
        return redirect(url_for("admin.structure"))
    if District.query.filter_by(name=name).first():
        flash("Такой район уже существует", "error")
        return redirect(url_for("admin.structure"))
    db.session.add(District(name=name))
    db.session.commit()
    flash("Район добавлен", "ok")
    return redirect(url_for("admin.structure"))


@bp.post("/mahalla/create")
@login_required
def mahalla_create():
    require_role(current_user, "admin")
    name = request.form.get("name", "").strip()
    district_id = request.form.get("district_id", type=int)
    if not name or not district_id:
        flash("Заполните все поля", "error")
        return redirect(url_for("admin.structure"))

    duplicate = Mahalla.query.filter(
        Mahalla.district_id == district_id,
        func.lower(Mahalla.name) == name.lower(),
    ).first()
    if duplicate:
        flash("Такая махалля уже есть в этом районе", "error")
        return redirect(url_for("admin.structure"))

    db.session.add(Mahalla(name=name, district_id=district_id))
    db.session.commit()
    flash("Махалля добавлена", "ok")
    return redirect(url_for("admin.structure"))


@bp.post("/street/create")
@login_required
def street_create():
    require_role(current_user, "admin")
    name = request.form.get("name", "").strip()
    mahalla_id = request.form.get("mahalla_id", type=int)
    if not name or not mahalla_id:
        flash("Заполните все поля", "error")
        return redirect(url_for("admin.structure"))

    duplicate = Street.query.filter(
        Street.mahalla_id == mahalla_id,
        func.lower(Street.name) == name.lower(),
    ).first()
    if duplicate:
        flash("Такая улица уже есть в этой махалле", "error")
        return redirect(url_for("admin.structure"))

    db.session.add(Street(name=name, mahalla_id=mahalla_id))
    db.session.commit()
    flash("Улица добавлена", "ok")
    return redirect(url_for("admin.structure"))


@bp.get("/workers")
@login_required
def workers():
    require_role(current_user, "admin")
    q = request.args.get("q", "").strip()
    page = request.args.get("page", default=1, type=int)
    per_page = request.args.get("per_page", default=100, type=int)
    per_page = max(20, min(per_page, 200))

    workers_q = User.query.filter_by(role="worker", is_active=True)
    if q:
        like = f"%{q}%"
        workers_q = workers_q.filter(or_(
            User.first_name.ilike(like),
            User.last_name.ilike(like),
            User.middle_name.ilike(like),
            User.login.ilike(like),
            (User.last_name + " " + User.first_name + " " + User.middle_name).ilike(like)
        ))
    workers_page = workers_q.order_by(User.last_name.asc()).paginate(page=page, per_page=per_page, error_out=False)
    workers = workers_page.items

    free_streets_query = (
        Street.query
        .outerjoin(WorkerAssignment, WorkerAssignment.street_id == Street.id)
        .filter(WorkerAssignment.id.is_(None))
    )
    free_streets_total = free_streets_query.count()
    free_streets = free_streets_query.order_by(Street.name.asc()).limit(500).all()

    return render_template(
        "admin_workers.html",
        workers=workers,
        workers_page=workers_page,
        per_page=per_page,
        free_streets=free_streets,
        free_streets_total=free_streets_total,
        q=q,
    )


@bp.post("/worker/create")
@login_required
def worker_create():
    require_role(current_user, "admin")

    street_id = request.form.get("street_id", type=int)
    login = request.form.get("login", "").strip()
    password = request.form.get("password", "").strip()
    first_name = request.form.get("first_name", "").strip()
    last_name = request.form.get("last_name", "").strip()
    middle_name = request.form.get("middle_name", "").strip()
    note = request.form.get("note", "").strip()

    if not all([street_id, login, password, first_name, last_name]):
        flash("Заполните обязательные поля", "error")
        return redirect(url_for("admin.workers"))

    if User.query.filter_by(login=login).first():
        flash("Логин занят", "error")
        return redirect(url_for("admin.workers"))

    if WorkerAssignment.query.filter_by(street_id=street_id).first():
        flash("На этой улице уже есть работник", "error")
        return redirect(url_for("admin.workers"))

    user = User(
        role="worker",
        login=login,
        password_hash=hash_password(password),
        first_name=first_name,
        last_name=last_name,
        middle_name=middle_name,
        note=note,
    )
    db.session.add(user)
    db.session.flush()

    db.session.add(WorkerAssignment(street_id=street_id, worker_id=user.id))
    db.session.commit()

    flash("Работник создан", "ok")
    return redirect(url_for("admin.workers"))


@bp.post("/worker/update/<int:worker_id>")
@login_required
def worker_update(worker_id: int):
    require_role(current_user, "admin")

    user = User.query.filter_by(id=worker_id, role="worker").first_or_404()

    user.first_name = request.form.get("first_name", user.first_name).strip()
    user.last_name = request.form.get("last_name", user.last_name).strip()
    user.middle_name = request.form.get("middle_name", user.middle_name).strip()
    user.note = request.form.get("note", user.note).strip()

    new_pass = request.form.get("password", "").strip()
    if new_pass:
        user.password_hash = hash_password(new_pass)

    db.session.add(user)
    db.session.commit()
    flash("Работник обновлён", "ok")
    return redirect(url_for("admin.workers"))


@bp.post("/worker/delete/<int:worker_id>")
@login_required
def worker_delete(worker_id: int):
    require_role(current_user, "admin")

    user = User.query.filter_by(id=worker_id, role="worker").first_or_404()

    a = WorkerAssignment.query.filter_by(worker_id=user.id).first()
    if a:
        db.session.delete(a)

    user.is_active = False
    user.login = f"{user.login}__deleted__{user.id}"
    db.session.add(user)
    db.session.commit()
    flash("Работник удалён", "ok")
    return redirect(url_for("admin.workers"))


@bp.post("/district/update/<int:district_id>")
@login_required
def district_update(district_id: int):
    require_role(current_user, "admin")
    d = District.query.get_or_404(district_id)
    name = request.form.get("name", "").strip()
    if not name:
        flash("Название района не может быть пустым", "error")
        return redirect(url_for("admin.structure"))
    exists = District.query.filter(District.name == name, District.id != d.id).first()
    if exists:
        flash("Район с таким названием уже существует", "error")
        return redirect(url_for("admin.structure"))
    d.name = name
    db.session.commit()
    flash("Район обновлён", "ok")
    return redirect(url_for("admin.structure"))


@bp.post("/district/delete/<int:district_id>")
@login_required
def district_delete(district_id: int):
    require_role(current_user, "admin")
    d = District.query.get_or_404(district_id)

    if Mahalla.query.filter_by(district_id=d.id).count() > 0:
        flash("Нельзя удалить район: в нём есть махалли", "error")
        return redirect(url_for("admin.structure"))

    db.session.delete(d)
    db.session.commit()
    flash("Район удалён", "ok")
    return redirect(url_for("admin.structure"))


@bp.post("/mahalla/update/<int:mahalla_id>")
@login_required
def mahalla_update(mahalla_id: int):
    require_role(current_user, "admin")
    m = Mahalla.query.get_or_404(mahalla_id)
    name = request.form.get("name", "").strip()
    district_id = request.form.get("district_id", type=int)

    if not name or not district_id:
        flash("Заполните название и район", "error")
        return redirect(url_for("admin.structure"))

    duplicate = Mahalla.query.filter(
        Mahalla.id != m.id,
        Mahalla.district_id == district_id,
        func.lower(Mahalla.name) == name.lower(),
    ).first()
    if duplicate:
        flash("Такая махалля уже есть в этом районе", "error")
        return redirect(url_for("admin.structure"))

    m.name = name
    m.district_id = district_id
    db.session.commit()
    flash("Махалля обновлена", "ok")
    return redirect(url_for("admin.structure"))


@bp.post("/mahalla/delete/<int:mahalla_id>")
@login_required
def mahalla_delete(mahalla_id: int):
    require_role(current_user, "admin")
    m = Mahalla.query.get_or_404(mahalla_id)

    if Street.query.filter_by(mahalla_id=m.id).count() > 0:
        flash("Нельзя удалить махаллю: в ней есть улицы", "error")
        return redirect(url_for("admin.structure"))

    db.session.delete(m)
    db.session.commit()
    flash("Махалля удалена", "ok")
    return redirect(url_for("admin.structure"))


@bp.post("/street/update/<int:street_id>")
@login_required
def street_update(street_id: int):
    require_role(current_user, "admin")
    s = Street.query.get_or_404(street_id)
    name = request.form.get("name", "").strip()
    mahalla_id = request.form.get("mahalla_id", type=int)

    if not name or not mahalla_id:
        flash("Заполните название и махаллю", "error")
        return redirect(url_for("admin.structure"))

    duplicate = Street.query.filter(
        Street.id != s.id,
        Street.mahalla_id == mahalla_id,
        func.lower(Street.name) == name.lower(),
    ).first()
    if duplicate:
        flash("Такая улица уже есть в этой махалле", "error")
        return redirect(url_for("admin.structure"))

    s.name = name
    s.mahalla_id = mahalla_id
    db.session.commit()
    flash("Улица обновлена", "ok")
    return redirect(url_for("admin.structure"))


@bp.post("/street/delete/<int:street_id>")
@login_required
def street_delete(street_id: int):
    require_role(current_user, "admin")
    s = Street.query.get_or_404(street_id)

    if WorkerAssignment.query.filter_by(street_id=s.id).first():
        flash("Нельзя удалить улицу: на ней уже есть работник", "error")
        return redirect(url_for("admin.structure"))

    db.session.delete(s)
    db.session.commit()
    flash("Улица удалена", "ok")
    return redirect(url_for("admin.structure"))
